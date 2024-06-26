from openpyxl import Workbook
import os
from openpyxl.utils import get_column_letter
# -*- coding: utf-8 -*-


from random import choice, choices, randint


class GenerateTimeTable:
    """
    This class generates a timetable using a genetic algorithm. It takes several
    parameters that are used to customize the generated timetable. The algorithm runs
    for a maximum number of generations, and the best timetable (the one with the
    highest fitness score) is returned at the end.
    """

    def __init__(
        self,
        classes=4,
        courses=4,
        slots=7,
        days=5,
        repeat=2,
        teachers=1,
        population_size=40,
        max_fitness=100,
        max_generations=50,
        subject_codes_to_names=None,
    ):
        """
        Initializes a new instance of the GenerateTimeTable class.

        Parameters:
        classes (int): The number of classes that the timetable should be generated
            for.
        courses (int): The number of courses that need to be scheduled.
        slots (int): The number of time slots available in a day.
        days (int): The number of days in a week.
        repeat (int or list): The number of times a course should be taught in a day,
            or a list specifying how many times each course should be taught in a day.
        teachers (int or list): The number of teachers available for each course, or a
            list specifying how many teachers are available for each course.
        population_size (int): The number of timetables that should be generated in
            each generation.
        max_fitness (int): The maximum fitness score that a timetable can have.
        max_generations (int): The maximum number of generations that the algorithm
            should run for.
        """
        self.classes = classes
        self.courses = courses
        self.slots = slots
        self.days = days
        self.repeat = repeat
        self.teachers = teachers
        self.population_size = population_size
        self.max_fitness = max_fitness
        self.max_generations = max_generations
        self.course_count = 0
        self.slot_count = 0
        self.day_count = 0
        self.class_count = 0
        self.course_bits = 0
        self.slot_bits = 0
        self.class_bits = 0
        self.total_slots = 0
        self.course_quota = []
        self.teacher_quota = []
        self.repeat_quota = []
        self.tables = []
        self.subject_codes_to_names = subject_codes_to_names
    def get_subject_name(self, subject_code):
        """
        Returns the subject name corresponding to the given subject code.
        """
        return self.subject_codes_to_names.get(subject_code, f"Subject {subject_code}")


    def initialize_genotype(self, no_courses, classes, slots, days, daily_rep,
                            teachers):
        """
        Initializes and stores important data relevant to the user-defined
        timetable(s)'s design in global variables so that they can be easily used
        multiple times throughout the program as per requirement.

        Course_bits, slot_bits and class bits are the lengths of binary string needed
        to represent them respectively. For example if course_count is 8, then the
        maximum course number will b.e 8 which requires 4 bits, hence course_bits will
        be equal to 4.

        Parameters:
        no_courses (int): The number of courses that need to be scheduled.
        classes (int): The number of classes that the timetable should be generated
            for.
        slots (int): The number of time slots available in a day.
        days (int): The number of days in a week.
        daily_rep (int or list): The number of times a course should be taught in a
            day, or a list specifying how many times each course should be taught in
            a day.
        teachers (int or list): The number of teachers available for each course, or a
            list specifying how many teachers are available for each course.

        Returns:
        A list containing the number of bits required to encode a course, slot, and a
            complete gene.
        """
        self.course_count = no_courses
        self.slot_count = slots
        self.day_count = days
        self.class_count = classes

        self.total_slots = self.slot_count * self.day_count
        self.course_bits = len(bin(self.course_count)) - 2
        self.slot_bits = len(bin(self.total_slots)) - 2
        self.class_bits = len(bin(self.course_count)) - 2

        self.calc_course_quota()

        if isinstance(daily_rep, int):
            self.repeat_quota = [daily_rep for _ in range(self.course_count)]
        elif isinstance(daily_rep[0],
                        int) and len(daily_rep) == self.course_count:
            self.repeat_quota = daily_rep
        else:
            raise ValueError("Invalid data supplied for daily repetitions.")

        self.repeat_quota = [
            self.repeat_quota[:] for _ in range(self.class_count)
        ]

        if isinstance(teachers, int):
            self.teacher_quota = [teachers] * self.course_count
        elif isinstance(teachers[0],
                        int) and len(teachers) == self.course_count:
            self.teacher_quota = teachers
        else:
            raise ValueError("Invalid data supplied for teachers.")

        return [
            self.course_bits,
            self.slot_bits,
            self.slot_count * self.day_count * self.class_count,
        ]

    def calc_course_quota(self):
        """
        This function calculates an array course_quota which stores the maximum allowed
            occurrence of a course/subject/module in a week/scheduled number of days.
        """
        q_max = self.total_slots // self.course_count
        if self.total_slots % self.course_count == 0:
            self.course_quota = [q_max for _ in range(self.course_count)]
        else:
            self.course_quota = [(q_max + 1) for _ in range(self.course_count)]

            extra_slots = (q_max + 1) * self.course_count - self.total_slots

            n = randint(1, self.course_count - extra_slots)
            for i in range(extra_slots):
                self.course_quota[n + i] -= 1

        self.course_quota = [
            self.course_quota[:] for _ in range(self.class_count)
        ]

    def encode_class(self):
        """
        The encode_class() function generates random binary strings whose integer
            values represent a course/module/subject

        Left padding of random binary strings with 0 is done to ensure each string is
            of same consistent length.
        """
        class_code = bin(randint(1, self.class_count))[2:]
        class_code = "0" * (self.class_bits - len(class_code)) + class_code
        return class_code

    def encode_slot(self):
        """
        The encode_slot() function generates random binary strings whose integer
            values represent slot number for a day.
        """
        slot_code = bin(randint(1, self.total_slots))[2:]
        slot_code = "0" * (self.slot_bits - len(slot_code)) + slot_code
        return slot_code

    def encode_course(self):
        """
        The encode_slot() function generates random binary strings whose integer
            values represents a course/module/subject.
        """
        course_code = bin(randint(1, self.course_count))[2:]
        course_code = "0" * (self.course_bits - len(course_code)) + course_code
        return course_code

    def generate_gene(self):
        """
        Generates a gene by encoding the course, class, and slot codes and
            concatenating them.

        Returns:
            str: The generated gene.

        Example:
            >>> obj = ClassName()
            >>> obj.generate_gene()
            'ENGL101M01'
        """
        course_code = self.encode_course()
        class_code = self.encode_class()
        slot_code = self.encode_slot()
        return course_code + slot_code + class_code

    def extract_slot_day(self, gene):
        """
        The class_slot is a cumulative class slot number, we calculate day number
            and slot number for that day for a gene using this class_slot number.
        """
        class_slot = int(
            gene[self.course_bits:self.course_bits + self.slot_bits], 2)
        slot_no = class_slot % self.slot_count
        day_no = class_slot // self.slot_count

        if slot_no == 0:
            slot_no = self.slot_count
            day_no -= 1
        return slot_no, day_no

    def calculate_fitness(self, gene):
        """
        This function determines fitness_score of a gene(course schedule) by checking
            few things -
        1)   If there already exists a course schedule for the same slot of the same
            or different class, fitness_score of the gene is decreased.
        2)   If the same course is scheduled for any of the adjacent slots in the same
            class, fitness_score of that gene is reduced.
        3)   If a course is occurring more han a fixed number of times, the
            fitness_score of that gene is reduced.
        """
        fitness_score = 100
        course = int(gene[0:self.course_bits], 2)

        slot_no, day_no = self.extract_slot_day(gene)
        class_no = int(gene[self.course_bits + self.slot_bits:], 2)

        if self.tables[class_no - 1][day_no - 1][slot_no - 1] != 0:
            fitness_score *= 0.01

        for i in range(self.class_count):
            if self.tables[i][day_no - 1][slot_no - 1] == course:
                fitness_score *= 0.6

        if (slot_no != 1 and self.tables[class_no - 1][day_no - 1][slot_no - 2]
                == course):
            fitness_score *= 0.6

        if (slot_no != self.slot_count
                and self.tables[class_no - 1][day_no - 1][slot_no] == course):
            fitness_score *= 0.6

        if self.course_quota[class_no - 1][course - 1] < 1:
            fitness_score *= 0.01

        if self.tables[class_no - 1][day_no - 1].count(course) >= 2:
            fitness_score *= 0.01

        if (self.tables[class_no - 1][day_no - 1].count(course) >=
                self.repeat_quota[class_no - 1][course - 1]):
            fitness_score *= 0.5

        temp_counter = 0
        for i in range(self.class_count):
            if self.tables[i][day_no - 1][slot_no - 1] == course:
                temp_counter += 1
        if temp_counter == self.teacher_quota[course - 1]:
            fitness_score *= 0.01
        return fitness_score

    def generate_table_skeleton(self):
        """
        This function returns a 3d array with 0 value for all positions. We use this
            array to store the schedules and the timetables.
        """
        for _ in range(self.class_count):
            class_table = []
            for _ in range(self.day_count):
                day = [0 for _ in range(self.slot_count)]
                class_table.append(day)
            self.tables.append(class_table)
        return self.tables

    def fit_slot(self, gene):
        """
        The fit_slot() function fills the tables array with fit course schedules that
            are returned by run_evolution().

        Python list indexing starts from 0, hence we subtract 1 from class_no, day_no,
        slot_no which are natural numbers.
        """
        course = int(gene[0:self.course_bits], 2)

        slot_no, day_no = self.extract_slot_day(gene)
        class_no = int(gene[self.course_bits + self.slot_bits:], 2)

        self.tables[class_no - 1][day_no - 1][slot_no - 1] = course
        self.course_quota[class_no - 1][course - 1] -= 1

    def generate_population(self, size):
        """
        Generates a population of genes by repeatedly calling the `generate_gene`
           method.

        Args:
            size (int): The desired size of the population.

        Returns:
            List[str]: A list of generated genes.

        Example:
            >>> obj = ClassName()
            >>> obj.generate_population(3)
            ['ENGL101M01', 'MATH102T05', 'PHYS201W03']
        """
        return [self.generate_gene() for _ in range(size)]

    def single_point_crossover(self, gene_a, gene_b):
        """
        For crossover, we randomly choose one out of course_code, slot_code and
            class_code to swap between the genes.
        """
        c = choice([1, 2, 3])

        if c == 1:
            gene_c = gene_b[0:self.course_bits] + gene_a[self.course_bits:]
            gene_d = gene_a[0:self.course_bits] + gene_b[self.course_bits:]

        elif c == 2:
            gene_c = (
                gene_a[:self.course_bits] +
                gene_b[self.course_bits:self.course_bits + self.slot_bits] +
                gene_a[self.course_bits + self.slot_bits:])
            gene_d = (
                gene_b[:self.course_bits] +
                gene_a[self.course_bits:self.course_bits + self.slot_bits] +
                gene_b[self.course_bits + self.slot_bits:])

        else:
            gene_c = (gene_a[:self.course_bits + self.slot_bits] +
                      gene_b[self.course_bits + self.slot_bits:])
            gene_d = (gene_b[:self.course_bits + self.slot_bits] +
                      gene_a[self.course_bits + self.slot_bits:])
        return [gene_c, gene_d]

    def multi_point_crossover(self, gene_a, gene_b, points):
        """
        Performs multi-point crossover on two input genes.

        Args:
            gene_a (str): The first gene to be crossed over.
            gene_b (str): The second gene to be crossed over.
            points (int): The number of crossover points to use.

        Returns:
            List[str]: A list containing the two offspring genes generated by
            multi-point crossover.

        Example:
            >>> obj = ClassName()
            >>> gene_a = 'ENGL101M01'
            >>> gene_b = 'MATH102T05'
            >>> obj.multi_point_crossover(gene_a, gene_b, 2)
            ['ENGL102T05', 'MATH101M01']
        """
        for _ in range(points):
            gene_a, gene_b = self.single_point_crossover(gene_a, gene_b)

        return [gene_a, gene_b]

    def mutation(self, gene, course_bit_length, slot_bit_length):
        """
        Applies mutation to the gene by randomly replacing one of its codes with a new
            random code of the same type.

        Parameters:
        gene (str): The gene to be mutated.
        course_bit_length (int): The length of the course code in bits.
        slot_bit_length (int): The length of the slot code in bits.

        Returns:
        str: The mutated gene.
        """
        c = choice([1, 2, 3])

        if c == 1:
            random_course = self.encode_course()
            mutated_gene = random_course + gene[course_bit_length:]

        elif c == 2:
            random_slot = self.encode_slot()
            mutated_gene = (gene[:course_bit_length] + random_slot +
                            gene[course_bit_length + slot_bit_length:])
        else:
            random_class = self.encode_class()
            mutated_gene = gene[:course_bit_length +
                                slot_bit_length] + random_class

        return mutated_gene

    def selection_pair(self, population):
        """
        Selects two individuals from the population, based on the fitness of each
            individual.

        Args:
            population (list): A list of individuals.

        Returns:
            A list of two individuals selected from the population.
        """
        return choices(
            population=population,
            weights=[self.calculate_fitness(gene) for gene in population],
            k=2,
        )

    def sort_population(self, population):
        """
        Sorts the population in descending order based on the fitness of each
            individual.

        Args:
            population (list): A list of individuals.

        Returns:
            The sorted population.
        """
        return sorted(population, key=self.calculate_fitness, reverse=True)

    def run_evolution(
        self,
        course_bit_length,
        slot_bit_length,
        population_size,
        max_fitness,
        max_generations,
    ):
        """
        Runs the evolutionary algorithm to generate the schedule.

        Args:
            course_bit_length (int): The length of the course bit string.
            slot_bit_length (int): The length of the slot bit string.
            population_size (int): The size of the population.
            max_fitness (int): The maximum fitness value to be achieved.
            max_generations (int): The maximum number of generations to be run.

        Returns:
            The best individual generated by the algorithm.
        """
        population = self.generate_population(population_size)
        for _ in range(max_generations):
            population = sorted(population,
                                key=self.calculate_fitness,
                                reverse=True)

            if self.calculate_fitness(population[0]) >= max_fitness:
                return population[0]

            next_generation = population[0:2]

            for _ in range(len(population) // 2 - 1):
                parents = self.selection_pair(population)
                children = self.single_point_crossover(parents[0], parents[1])
                child_a = self.mutation(children[0], course_bit_length,
                                        slot_bit_length)
                child_b = self.mutation(children[1], course_bit_length,
                                        slot_bit_length)
                next_generation += [child_a, child_b]

            population = next_generation
        return population[0]

    def run(self):
        """
        Runs the scheduling algorithm.

        Returns:
            The generated schedule.
        """
        course_bit_length, slot_bit_length, all_slots = self.initialize_genotype(
            self.courses,
            self.classes,
            self.slots,
            self.days,
            self.repeat,
            self.teachers,
        )
        self.generate_table_skeleton()
        while all_slots > 0:
            gene = self.run_evolution(
                course_bit_length,
                slot_bit_length,
                self.population_size,
                self.max_fitness,
                self.max_generations,
            )
            if gene != 0:
                self.fit_slot(gene)
                all_slots -= 1
        return self.tables
        


total_classes = 4
no_courses = 5
slots = 6
total_days = 5
daily_repetition = 2
teachers = [1,2,3,2,2]

my_dict = {
    0: "FOC",
    1: "DBM",
    2: "AD",
    3: "DM",
    4: "INE",
    5: "FOC"
}
my_week = {
    0: "Mon",
    1: "Tue",
    2: "Wed",
    3: "Thur",
    4: "Fri"

}

table = GenerateTimeTable(total_classes, no_courses, slots, total_days, daily_repetition,teachers)
i=1
j=0
count = 5
for single_table in table.run():
    print("BE "+str(count))
    count = count+ 1
    i=i+1
    for days in single_table:
       print(my_week[j]+str("->"),end=' ')
       j=j+1
       for item in days:

         print(my_dict[item], end=' ')
       print("")
    j=0





    print("-----------------------------------")

# Your existing code goes here...

# Initialize a new workbook
file_path = "timetable.xlsx"

# Check if the file exists and delete it if it does
if os.path.exists(file_path):
    os.remove(file_path)

# Initialize a new workbook
wb = Workbook()

# Function to convert slot codes to subject names
def get_subject_name(subject_code):
    return my_dict[subject_code]

# Function to convert day codes to day names
def get_day_name(day_code):
    return my_week[day_code]

# Function to convert time to 12-hour format
def convert_to_12_hour_format(hour):
    if hour == 0:
        return "12 AM"
    elif hour < 12:
        return f"{hour} AM"
    elif hour == 12:
        return "12 PM"
    else:
        return f"{hour - 12} PM"

# Write timetable data to Excel for each class
for class_idx, class_table in enumerate(table.run(), start=5):
    class_name = f"Be{class_idx}"
    ws = wb.create_sheet(title=f"{class_name} Timetable")
    ws.cell(row=1, column=1).value = f"{class_name} Timetable"
    ws.cell(row=2, column=1).value = "Day/Time"
    time = 8  # Starting time
    for slot_idx, day_table in enumerate(class_table, start=2):
        ws.cell(row=2, column=slot_idx).value = f"{convert_to_12_hour_format(time)} - {convert_to_12_hour_format(time + 1)}"
        time += 1
        # If it's the 6th slot, break the loop
        if slot_idx == 7:
            break
    ws.cell(row=2, column=slot_idx + 1).value = f"{convert_to_12_hour_format(time)} - {convert_to_12_hour_format(time + 1)}"

    for day_idx, day_table in enumerate(class_table, start=3):
        ws.cell(row=day_idx, column=1).value = get_day_name(day_idx-3)
        time = 8  # Reset starting time
        for slot_idx, slot in enumerate(day_table, start=2):
            ws.cell(row=day_idx, column=slot_idx).value = get_subject_name(slot)
            time += 1
        # Add the 6th slot for each day

        
for col in ws.columns:
    max_length = 0
    column = col[0].column
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width


# Save the workbook# Save the workbook
wb.save(file_path)